mport os
import sys
import uvicorn
import re
import xml.etree.ElementTree as ET
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
import pdfplumber  
import pandas as pd
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_chroma import Chroma
from langchain_groq import ChatGroq
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_classic.chains import create_retrieval_chain  
from langchain_classic.chains.combine_documents import create_stuff_documents_chain  
from langchain_core.prompts import ChatPromptTemplate
from dotenv import load_dotenv

# ==========================================
# 🔑 CONFIGURACIÓN SEGURA DE ENTORNO
# ==========================================
load_dotenv()
GROQ_API_KEY = os.getenv("GROQ_API_KEY")

if not GROQ_API_KEY:
    # 📑 REEMPLAZÁ ESTA CADENA POR TU NUEVA API KEY ACTIVA GENERADA EN GROQ:
    os.environ["GROQ_API_KEY"] = "AcÁ va tu clave groq"
else:
    os.environ["GROQ_API_KEY"] = GROQ_API_KEY


CARPETA_PDFS = r"C:\Users\Usuario\Desktop\automatizar\mi IA\mis_pdfs" 
CARPETA_VECTORIAL = "./base_datos_chroma_multi"

# Variables globales unificadas para el ciclo de vida de FastAPI e i3
base_datos = None
cadena_documentos = None
ultima_respuesta_ia = "" 
archivos_disponibles = []
ultimas_mediciones_fisicas = {"humedad_suelo": 0.0, "temperatura": 0.0}

class Consulta(BaseModel):
    pregunta: str

class DatosSensores(BaseModel):
    humedad_suelo: float
    temperatura: float

# ==========================================
# 📑 PROCESADORES DE ESTRUCTURAS Y ARCHIVOS
# ==========================================
def convertir_tabla_a_markdown(tabla):
    """Transforma una matriz de celdas de pdfplumber en una tabla Markdown limpia."""
    if not tabla or not any(tabla):
        return ""
    
    lineas = []
    tabla_limpia = [[str(celda).replace('\n', ' ').strip() if celda is not None else "" for celda in fila] for fila in tabla]
    
    lineas.append("| " + " | ".join(tabla_limpia) + " |")
    lineas.append("| " + " | ".join(["---"] * len(tabla_limpia)) + " |")
    
    for fila in tabla_limpia[1:]:
        if any(fila):  
            lineas.append("| " + " | ".join(fila) + " |")
            
    return "\n".join(lineas) + "\n\n"

def extraer_texto_de_xml(ruta_xml):
    """Extrae todo el texto visible de un archivo XML/XLM de forma segura."""
    try:
        tree = ET.parse(ruta_xml)
        root = tree.getroot()
        texto_completo = "".join(root.itertext())
        return texto_completo.strip()
    except Exception as e:
        print(f"❌ Error al leer el archivo XML {ruta_xml}: {e}")
        return ""

def cargar_multiples_documentos(ruta_carpeta):
    """Escanea la carpeta extrayendo texto plano y tablas estructuradas (PDF y XML)."""
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)
        return []
    
    documentos_langchain = []
    global archivos_disponibles
    archivos_disponibles = os.listdir(ruta_carpeta)
    
    for nombre_archivo in archivos_disponibles:
        ruta_completa = os.path.join(ruta_carpeta, nombre_archivo)
        ext = nombre_archivo.lower()
        
        if ext.endswith('.pdf'):
            try:
                with pdfplumber.open(ruta_completa) as pdf:
                    for numero_pagina, pagina in enumerate(pdf.pages):
                        tablas_visibles = pagina.extract_tables()
                        texto_tablas_md = ""
                        for tabla in tablas_visibles:
                            texto_tablas_md += convertir_tabla_a_markdown(tabla)
                        
                        texto_plano = pagina.extract_text() or ""
                        
                        contenido_final = ""
                        if texto_tablas_md:
                            contenido_final += f"--- ESTRUCTURA DE TABLA DETECTADA ---\n{texto_tablas_md}"
                        if texto_plano.strip():
                            contenido_final += f"--- TEXTO DE LA PÁGINA ---\n{texto_plano}"
                            
                        if contenido_final.strip():
                            doc = Document(
                                page_content=contenido_final,
                                metadata={"source": str(nombre_archivo), "page": int(numero_pagina + 1)}
                            )
                            documentos_langchain.append(doc)
            except Exception as e:
                print(f"❌ Error en PDF {nombre_archivo}: {e}")
                continue
                
        elif ext.endswith('.xml') or ext.endswith('.xlm'):
            texto_xml = extraer_texto_de_xml(ruta_completa)
            if texto_xml:
                doc = Document(
                    page_content=texto_xml,
                    metadata={"source": str(nombre_archivo), "page": 1}
                )
                documentos_langchain.append(doc)
                
    return documentos_langchain
# ==========================================
# 🧠 INICIALIZACIÓN DEL SISTEMA VECTORIAL
# ==========================================
def inicializar_sistema():
    """Genera los vectores en el disco y conecta la arquitectura con Groq."""
    global base_datos, cadena_documentos, archivos_disponibles
    print("🧠 Inicializando modelo de lenguaje para procesamiento de texto...")
    
    embeddings = HuggingFaceEmbeddings(
        model_name="sentence-transformers/all-MiniLM-L6-v2",
        cache_folder="./modelo_embeddings_cache"
    )
    
    if os.path.exists(CARPETA_PDFS):
        archivos_disponibles = os.listdir(CARPETA_PDFS)
    
    if os.path.exists(CARPETA_VECTORIAL):
        print("📚 Reutilizando base de datos híbrida indexada en el disco...")
        base_datos = Chroma(persist_directory=CARPETA_VECTORIAL, embedding_function=embeddings)
    else:
        print("⏳ Procesando tus archivos PDF y XML por primera vez...")
        documentos = cargar_multiples_documentos(CARPETA_PDFS)
        if not documentos:
            base_datos = Chroma.from_documents(
                documents=[Document(page_content="Carpeta vacía", metadata={"source": "vacio"})], 
                embedding=embeddings, 
                persist_directory=CARPETA_VECTORIAL
            )
        else:
            separador = RecursiveCharacterTextSplitter(chunk_size=1800, chunk_overlap=400)
            fragmentos = separador.split_documents(documentos)
            base_datos = Chroma.from_documents(
                documents=fragmentos, 
                embedding=embeddings, 
                persist_directory=CARPETA_VECTORIAL
            )
        print("✅ ¡Documentos indexados con éxito!")
    
    llm = ChatGroq(model="llama-3.1-8b-instant", temperature=0.0)
    
    instrucciones = (
        "Sos un asistente de IA expert en análisis técnico de documentos (PDFs y XMLs).\n"
        "Tu único objetivo es responder la pregunta del usuario utilizando de forma estricta "
        "y exclusiva el contexto provisto abajo. No uses conocimientos externos.\n\n"
        
        "REGLAS OBLIGATORIAS DE RESPUESTA (REQUERIMIENTO DE INTERFAZ):\n"
        "1. IDIOMA Y TONO: Responde siempre en español de Argentina de forma clara y directa.\n"
        "2. ESTRUCTURA DE TABLAS (CRUCIAL): Si los datos del contexto contienen matrices, cuadros técnicos o dosificaciones, DEBES armar obligatoriamente tablas usando el formato Markdown clásico con barras verticales.\n"
        "Ejemplo exacto de formato obligatorio:\n"
        "| Clase de Herbicida | Efecto | Uso |\n"
        "| --- | --- | --- |\n"
        "| Selectivos | Matan solo malezas | Cultivos |\n"
        "Usa el formato tabular siempre que sea lógicamente posible ordenar la información en filas y columnas.\n"
        "3. TRATAMIENTO DE LA DUDA: Si la respuesta no figura explícitamente en el contexto, di exactamente: "
        "'Lo siento, no encuentro esa información en los documentos.' No inventes nada bajo ninguna circunstancia.\n\n"
        "Contexto:\n{context}"
    )
    prompt = ChatPromptTemplate.from_messages([("system", instrucciones), ("human", "{input}")])
    cadena_documentos = create_stuff_documents_chain(llm, prompt)
    print("🚀 Servidor de IA indexado y listo.")

@asynccontextmanager
async def lifespan(app: FastAPI):
    inicializar_sistema()
    yield

app = FastAPI(title="API Servidor RAG", lifespan=lifespan)

@app.get("/")
def ruta_raiz():
    return {"status": "online"}

@app.post("/actualizar-sensores")
def actualizar_sensores(datos: DatosSensores):
    global ultimas_mediciones_fisicas
    ultimas_mediciones_fisicas["humedad_suelo"] = datos.humedad_suelo
    ultimas_mediciones_fisicas["temperatura"] = datos.temperatura
    return {"status": "telemetria_recibida", "datos": ultimas_mediciones_fisicas}

@app.post("/preguntar")
def preguntar_ia(consulta: Consulta):
    global base_datos, cadena_documentos, ultima_respuesta_ia, archivos_disponibles, ultimas_mediciones_fisicas
    if base_datos is None or cadena_documentos is None:
        raise HTTPException(status_code=503, detail="El sistema de IA no está inicializado.")
    try:
        if not archivos_disponibles and os.path.exists(CARPETA_PDFS):
            archivos_disponibles = os.listdir(CARPETA_PDFS)

               # 🕵️‍♂️ ENRUTADOR DINÁMICO DE ARCHIVOS (CORREGIDO)
        archivo_objetivo = None
        for archivo in archivos_disponibles:
            # os.path.splitext devuelve una tupla (nombre, ext). Usamos [0] para sacar solo el nombre.
            nombre_sin_ext = os.path.splitext(archivo)[0] 
            if nombre_sin_ext.lower() in consulta.pregunta.lower() or archivo.lower() in consulta.pregunta.lower():
                archivo_objetivo = archivo
                break

        
        search_kwargs = {"k": 7}
        if archivo_objetivo:
            search_kwargs["filter"] = {"source": archivo_objetivo}
            print(f"🎯 Consulta enrutada exclusivamente al archivo: {archivo_objetivo}")
        else:
            print("🌍 Consulta global distribuida en todos los documentos.")
            
        recuperador_dinamico = base_datos.as_retriever(search_kwargs=search_kwargs)
        sistema_ia_dinamico = create_retrieval_chain(recuperador_dinamico, cadena_documentos)
        
        bloque_sensores = (
            f"\n\n[DATOS DE SENSORES EN TIEMPO REAL EN EL CAMPO]:\n"
            f"- Humedad actual del suelo: {ultimas_mediciones_fisicas['humedad_suelo']}% \n"
            f"- Temperatura ambiente: {ultimas_mediciones_fisicas['temperatura']}°C\n"
            f"Utilizá estos valores físicos actuales para contrastarlos con los límites técnicos de los PDFs.\n"
        )
        
        pregunta_enriquecida = consulta.pregunta + bloque_sensores
        resultado = sistema_ia_dinamico.invoke({"input": pregunta_enriquecida})
        ultima_respuesta_ia = resultado["answer"]
        
        fuentes = []
        documentos_recuperados = resultado.get("context", [])
            
        for doc in documentos_recuperados:
            archivo = doc.metadata.get("source", "Desconocido")
            pagina = doc.metadata.get("page", 1)
            
            if archivo and archivo != "vacio" and archivo != "Desconocido":
                if archivo.lower().endswith(('.xml', '.xlm')):
                    fuentes.append(f"• {archivo}")
                else:
                    fuentes.append(f"• {archivo} (Pág. {pagina})")
                    
        fuentes_unicas = list(sorted(set(fuentes)))
        return {"respuesta": resultado["answer"], "fuentes": fuentes_unicas}
    except Exception as e:
        import traceback
        print("❌ ERROR DETECTADO EN EL ENDPOINT /PREGUNTAR:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
@app.post("/exportar")
def exportar_a_excel():
    global ultima_respuesta_ia
    if not ultima_respuesta_ia:
        raise HTTPException(status_code=400, detail="No hay datos previos para exportar.")
    try:
        lineas = ultima_respuesta_ia.split('\n')
        tablas_encontradas = []
        tabla_actual = []
        
        for linea in lineas:
            if '|' in linea:
                if re.match(r'^\s*\|[\s\-:|]*\|\s*$', linea):
                    continue
                celdas = [c.strip() for c in linea.split('|')[1:-1]]
                if celdas:
                    tabla_actual.append(celdas)
            else:
                if tabla_actual:
                    tablas_encontradas.append(tabla_actual)
                    tabla_actual = []
        if tabla_actual:
            tablas_encontradas.append(tabla_actual)
            
        if not tablas_encontradas:
            raise HTTPException(status_code=404, detail="No se encontraron tablas.")
            
                # --- REPARACIÓN DE VARIABLES MATRICIALES EN EXPORTAR ---
        datos_tabla = tablas_encontradas[0] # Extraemos la primera tabla detectada
        if len(datos_tabla) < 2:
            raise HTTPException(status_code=404, detail="La tabla no contiene datos suficientes.")
            
        encabezados = datos_tabla[0] # La primera fila son los títulos
        filas = datos_tabla[1:]      # El resto son los registros

        
        num_columnas = len(encabezados)
        filas_normalizadas = []
        for f in filas:
            if len(f) < num_columnas:
                f = f + [""] * (num_columnas - len(f))
            elif len(f) > num_columnas:
                f = f[:num_columnas]
            filas_normalizadas.append(f)
        
        df = pd.DataFrame(filas_normalizadas, columns=encabezados)
        ruta_escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
        ruta_excel = os.path.join(ruta_escritorio, 'Tabla_Exportada_IA.xlsx')
        
        with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Datos IA")
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            workbook = writer.book
            worksheet = writer.sheets["Datos IA"]
            
            fill_cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font_cabecera = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            borde_fino = Side(border_style="thin", color="D9D9D9")
            cuadrícula = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
            
            for cell in worksheet:
                cell.fill = fill_cabecera
                cell.font = font_cabecera
                cell.alignment = align_centro
                cell.border = cuadrícula
            
            for col_idx, col in enumerate(worksheet.columns, start=1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                
                for cell in col:
                    if cell.row > 1: 
                        cell.border = cuadrícula
                        cell.alignment = Alignment(vertical="center")
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                        
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        return {"status": "success", "archivo": ruta_excel}
    except Exception as e:
        import traceback
        print("❌ ERROR DETECTADO EN EL ENDPOINT /EXPORTAR:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)

